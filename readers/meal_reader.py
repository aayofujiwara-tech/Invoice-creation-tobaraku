"""食費管理表パーサー

食費管理表Excelファイルから、指定月の入居者ごとの食事提供データを読み取る。

構造:
  Row 1: A1=年, B1=月
  Row 2: ヘッダー
  Row 3: 曜日行
  Row 4+: 入居者データ（3行1セット: 朝食/昼食/夕食）
  列A,B,C: 3行結合（居室番号, 利用者名, 食事形態）
  列D: 食事種別（朝食/昼食/夕食）
  列E〜AI: 日付1〜31のフラグ（1.0=提供あり）
  列AJ: 食数合計
  列AK: 金額小計（朝食=330×食数, 昼食/夕食=550×食数）
  列AL: 請求金額（夕食行のみ、3食合計）
"""

from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

from utils.helpers import col_letter_to_index, normalize_room, to_reiwa_label


@dataclass
class MealRecord:
    """1入居者の月間食事データ"""
    room: str                    # 居室番号（正規化済み）
    name: str                    # 利用者名
    meal_form: str               # 食事形態
    breakfast_days: list[int] = field(default_factory=list)  # 朝食提供日のリスト (1-indexed)
    lunch_days: list[int] = field(default_factory=list)      # 昼食提供日のリスト
    dinner_days: list[int] = field(default_factory=list)     # 夕食提供日のリスト
    breakfast_count: int = 0
    lunch_count: int = 0
    dinner_count: int = 0
    excel_billing: int | None = None  # Excel上のAL列の値（検証用）

    @property
    def total_count(self) -> int:
        return self.breakfast_count + self.lunch_count + self.dinner_count

    def calc_billing(self, breakfast_price: int = 330,
                     lunch_price: int = 550,
                     dinner_price: int = 550) -> int:
        """食費月額を計算"""
        return (self.breakfast_count * breakfast_price
                + self.lunch_count * lunch_price
                + self.dinner_count * dinner_price)

    @property
    def is_empty(self) -> bool:
        """食事データが全くない（空室等）"""
        return self.total_count == 0


# 日付データの列範囲
DAY_COL_START = col_letter_to_index("E")   # 5
DAY_COL_END = col_letter_to_index("AI")    # 35
COUNT_COL = col_letter_to_index("AJ")      # 36
SUBTOTAL_COL = col_letter_to_index("AK")   # 37
BILLING_COL = col_letter_to_index("AL")    # 38


def _get_merged_value(ws, row: int, col: int):
    """結合セルの値を取得する。結合範囲の場合、左上セルの値を返す。"""
    cell = ws.cell(row=row, column=col)
    for merge_range in ws.merged_cells.ranges:
        if cell.coordinate in merge_range:
            return ws.cell(row=merge_range.min_row, column=merge_range.min_col).value
    return cell.value


def _read_day_flags(ws, row: int) -> tuple[list[int], int]:
    """1行分の日別フラグを読み取り、提供日リストと食数を返す。"""
    days = []
    count = 0
    for col in range(DAY_COL_START, DAY_COL_END + 1):
        day_num = col - DAY_COL_START + 1  # 1-indexed day
        val = ws.cell(row=row, column=col).value
        if val is not None and val == 1.0:
            days.append(day_num)
            count += 1
        elif val is not None and isinstance(val, (int, float)) and val == 1:
            days.append(day_num)
            count += 1
    return days, count


def read_meal_sheet(ws) -> list[MealRecord]:
    """食費管理表の1シートを解析して、入居者ごとのMealRecordリストを返す。

    Args:
        ws: openpyxlのワークシート

    Returns:
        MealRecordのリスト（空室含む）
    """
    records = []

    # データ開始行（Row 4から3行1セット）
    data_start = 4
    max_row = ws.max_row

    row = data_start
    while row + 2 <= max_row:
        # 3行1セットを読む
        # 列D で朝食/昼食/夕食を確認
        meal_type_1 = ws.cell(row=row, column=4).value
        meal_type_2 = ws.cell(row=row + 1, column=4).value
        meal_type_3 = ws.cell(row=row + 2, column=4).value

        # 朝食/昼食/夕食の3行パターンを確認
        if not (meal_type_1 and meal_type_2 and meal_type_3):
            # D列が空の場合はデータ終了
            break

        mt1 = str(meal_type_1).strip()
        mt2 = str(meal_type_2).strip()
        mt3 = str(meal_type_3).strip()

        if mt1 != "朝食" or mt2 != "昼食" or mt3 != "夕食":
            # 3行パターンに合致しない場合はスキップ
            row += 3
            continue

        # 居室番号・利用者名・食事形態（結合セルから取得）
        room_raw = _get_merged_value(ws, row, 1)
        name_raw = _get_merged_value(ws, row, 2)
        meal_form_raw = _get_merged_value(ws, row, 3)

        room = normalize_room(room_raw)
        name = str(name_raw).strip() if name_raw else ""
        meal_form = str(meal_form_raw).strip() if meal_form_raw else ""

        # 日別フラグを読み取り
        breakfast_days, breakfast_count = _read_day_flags(ws, row)
        lunch_days, lunch_count = _read_day_flags(ws, row + 1)
        dinner_days, dinner_count = _read_day_flags(ws, row + 2)

        # AL列（請求金額、夕食行のみ）
        billing_raw = ws.cell(row=row + 2, column=BILLING_COL).value
        excel_billing = int(billing_raw) if billing_raw else None

        record = MealRecord(
            room=room,
            name=name,
            meal_form=meal_form,
            breakfast_days=breakfast_days,
            lunch_days=lunch_days,
            dinner_days=dinner_days,
            breakfast_count=breakfast_count,
            lunch_count=lunch_count,
            dinner_count=dinner_count,
            excel_billing=excel_billing,
        )
        records.append(record)
        row += 3

    return records


def read_meal_file(filepath: str | Path, year: int, month: int) -> list[MealRecord]:
    """食費管理表ファイルから指定月のデータを読み取る。

    Args:
        filepath: 食費管理表Excelファイルのパス
        year: 西暦年（例: 2026）
        month: 月（例: 1）

    Returns:
        MealRecordのリスト
    """
    sheet_name = to_reiwa_label(year, month)
    wb = openpyxl.load_workbook(str(filepath), data_only=True)

    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(
            f"Sheet '{sheet_name}' not found in {filepath}. "
            f"Available: {wb.sheetnames}"
        )

    ws = wb[sheet_name]

    # Row 1 の年月を検証
    year_cell = ws.cell(row=1, column=1).value
    month_cell = ws.cell(row=1, column=2).value
    if year_cell is not None:
        file_year = int(year_cell)
        file_month = int(month_cell) if month_cell else 0
        if file_year != year or file_month != month:
            wb.close()
            raise ValueError(
                f"Year/month mismatch: expected {year}/{month}, "
                f"got {file_year}/{file_month} in {filepath} sheet {sheet_name}"
            )

    records = read_meal_sheet(ws)
    wb.close()
    return records


def read_all_facilities_meals(
    input_dir: str | Path,
    config: dict,
    year: int,
    month: int,
) -> dict[str, list[MealRecord]]:
    """全拠点の食費管理表を読み取る。

    Returns:
        {拠点名: [MealRecord, ...]} の辞書
    """
    input_path = Path(input_dir)
    result = {}

    for facility_name, fconf in config["facilities"].items():
        prefix = fconf["meal_file_prefix"]
        # ファイル名を検索
        candidates = list(input_path.glob(f"{prefix}*"))
        if not candidates:
            raise FileNotFoundError(
                f"Meal file not found for {facility_name}: {prefix}*"
            )
        filepath = candidates[0]
        records = read_meal_file(filepath, year, month)
        # 空でない入居者のみフィルタ（空室は除外）
        result[facility_name] = records

    return result
