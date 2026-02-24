"""領収書Excel生成モジュール

既存テンプレートのレイアウトを再現して、入居者ごとの領収書Excelを生成する。

テンプレート構造:
  B1     "領　収　書"
  B4     利用者名
  D4     "様"
  C6     金額（ご請求金額と同額）
  B7     "但し　家賃、共益費、管理費、水道代、自費分として"
  C9     "株式会社AA"
  C10    "〒555-0022"
  C11    "大阪市西淀川区柏里2-9-3-102"
  C12    "TEL・FAX 06-4400-2901"
"""

from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

from calc.billing import BillingResult
from utils.helpers import to_reiwa_label


def write_receipt(
    billing: BillingResult,
    year: int,
    month: int,
    issue_date: date,
    output_path: Path,
) -> Path:
    """1入居者分の領収書Excelファイルを生成する。

    Args:
        billing: BillingResult
        year: 対象年
        month: 対象月
        issue_date: 発行日
        output_path: 出力先パス

    Returns:
        保存先Path
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "領収書"

    # フォント定義
    title_font = Font(name="游ゴシック", bold=True, size=20)
    name_font = Font(name="游ゴシック", bold=True, size=14)
    normal_font = Font(name="游ゴシック", size=10)
    amount_font = Font(name="游ゴシック", bold=True, size=16)
    small_font = Font(name="游ゴシック", size=9)

    double_bottom = Border(bottom=Side(style="double"))

    # --- Row 1: タイトル ---
    ws.merge_cells("B1:D1")
    ws["B1"].value = "領　収　書"
    ws["B1"].font = title_font
    ws["B1"].alignment = Alignment(horizontal="center")

    # --- Row 2: 発行日 ---
    ws.merge_cells("D2:E2")
    ws["D2"].value = issue_date
    ws["D2"].number_format = "YYYY年MM月DD日"
    ws["D2"].font = small_font
    ws["D2"].alignment = Alignment(horizontal="right")

    # --- Row 4: 利用者名 ---
    ws.merge_cells("B4:C4")
    ws["B4"].value = billing.name
    ws["B4"].font = name_font
    ws["B4"].border = Border(bottom=Side(style="thin"))
    ws["C4"].border = Border(bottom=Side(style="thin"))
    ws["D4"].value = "様"
    ws["D4"].font = normal_font

    # --- Row 6: 金額 ---
    ws["B6"].value = "金額"
    ws["B6"].font = normal_font
    ws.merge_cells("C6:D6")
    ws["C6"].value = billing.total
    ws["C6"].font = amount_font
    ws["C6"].number_format = "¥#,##0"
    ws["C6"].alignment = Alignment(horizontal="center")
    ws["C6"].border = double_bottom
    ws["D6"].border = double_bottom

    # --- Row 7: 但し書き ---
    ws.merge_cells("B7:E7")
    ws["B7"].value = "但し　家賃、共益費、管理費、水道代、自費分として"
    ws["B7"].font = small_font

    # --- Row 9-12: 会社情報 ---
    ws.merge_cells("C9:E9")
    ws["C9"].value = "株式会社AA"
    ws["C9"].font = Font(name="游ゴシック", bold=True, size=11)

    ws.merge_cells("C10:E10")
    ws["C10"].value = "〒555-0022"
    ws["C10"].font = small_font

    ws.merge_cells("C11:E11")
    ws["C11"].value = "大阪市西淀川区柏里2-9-3-102"
    ws["C11"].font = small_font

    ws.merge_cells("C12:E12")
    ws["C12"].value = "TEL・FAX 06-4400-2901"
    ws["C12"].font = small_font

    # --- 列幅設定 ---
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 12

    # --- 印刷設定 ---
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.print_area = "A1:E14"

    # 保存
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    wb.close()

    return output_path


def write_all_receipts(
    billings: list[BillingResult],
    year: int,
    month: int,
    issue_date: date,
    output_dir: Path,
) -> list[Path]:
    """全入居者分の領収書を生成する。

    Returns:
        生成されたファイルパスのリスト
    """
    reiwa_label = to_reiwa_label(year, month)
    files = []

    for billing in billings:
        if billing.is_vacant:
            continue

        filename = f"領収書_{billing.name}_{reiwa_label}.xlsx"
        filepath = output_dir / filename

        write_receipt(
            billing=billing,
            year=year,
            month=month,
            issue_date=issue_date,
            output_path=filepath,
        )
        files.append(filepath)

    return files
