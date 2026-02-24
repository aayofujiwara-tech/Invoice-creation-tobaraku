"""請求マスター読取モジュール

ええすまい請求Excelファイル（各拠点）から、以下のデータを読み取る：
  1. 入居者マスタ（居室番号⇔利用者名）
  2. 月別集約シート（RX.X）の全入居者データ
  3. 請求書テンプレートの構造情報

RX.Xシート構造（新フォーマット R7.8以降）:
  Row 1: A1:B1結合 = 対象月末日(datetime)
  Row 2: ヘッダー行（A〜AC）
  Row 3+: 入居者データ（居室番号順、空室含む）
  ※空行は居室番号のみ or 全セル空白

列マッピング:
  A:居室 B:利用者名 C:分納残 D:家賃 E:管理費 F:共益費 G:水道料金
  H:定額光熱費 I:食事 J:調整額 K:オムツ L:日用品 M:分割支払い
  N:事務手数料 O:デイサービス P:福祉用具 Q:しろくま薬局
  R:往診診療費 S:サポート費 T:その他 U-V:(予備)
  W:計 X:一部負担(介護) Y:一部負担(看護) Z:合計
  AA:備考 AB:請求書 AC:分割残高
"""

from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

from utils.helpers import col_letter_to_index, normalize_room, to_reiwa_label


# RX.Xシートの列インデックス（1始まり）
COL = {
    "room": 1,            # A
    "name": 2,            # B
    "installment_bal": 3, # C 分納残
    "rent": 4,            # D 家賃
    "management": 5,      # E 管理費
    "common": 6,          # F 共益費
    "water": 7,           # G 水道料金
    "utility": 8,         # H 定額光熱費
    "meal": 9,            # I 食事
    "adjustment": 10,     # J 調整額
    "diaper": 11,         # K オムツ
    "daily_supplies": 12, # L 日用品
    "installment": 13,    # M 分割支払い
    "office_fee": 14,     # N 事務手数料
    "day_service": 15,    # O デイサービス
    "welfare_equip": 16,  # P 福祉用具
    "pharmacy": 17,       # Q しろくま薬局
    "doctor": 18,         # R 往診診療費
    "support": 19,        # S サポート費
    "other": 20,          # T その他
    "subtotal": 23,       # W 計
    "care_burden": 24,    # X 一部負担(介護)
    "nurse_burden": 25,   # Y 一部負担(看護)
    "total": 26,          # Z 合計
    "notes": 27,          # AA 備考
    "invoice_status": 28, # AB 請求書
    "remaining_bal": 29,  # AC 分割残高
}


@dataclass
class ResidentMaster:
    """入居者マスタの1行"""
    room: str
    name: str


@dataclass
class RxRow:
    """RX.Xシートの1行（1入居者分の請求データ）"""
    room: str
    name: str
    installment_balance: int | None = None  # C: 分納残
    rent: int | None = None                 # D: 家賃
    management: int | None = None           # E: 管理費
    common: int | None = None               # F: 共益費
    water: int | None = None                # G: 水道料金
    utility: int | None = None              # H: 定額光熱費
    meal: int | None = None                 # I: 食事
    adjustment: int | None = None           # J: 調整額
    diaper: int | None = None               # K: オムツ
    daily_supplies: int | None = None       # L: 日用品
    installment: int | None = None          # M: 分割支払い
    office_fee: int | None = None           # N: 事務手数料
    day_service: int | None = None          # O: デイサービス
    welfare_equip: int | None = None        # P: 福祉用具
    pharmacy: int | None = None             # Q: しろくま薬局
    doctor: int | None = None               # R: 往診診療費
    support: int | None = None              # S: サポート費
    other: int | None = None                # T: その他
    subtotal: int | None = None             # W: 計
    care_burden: int | None = None          # X: 一部負担(介護)
    nurse_burden: int | None = None         # Y: 一部負担(看護)
    total: int | None = None                # Z: 合計
    notes: str | None = None                # AA: 備考
    invoice_status: str | None = None       # AB: 請求書
    remaining_balance: int | None = None    # AC: 分割残高

    @property
    def is_vacant(self) -> bool:
        """空室かどうか"""
        return not self.name or self.total in (None, 0)

    @property
    def has_installment(self) -> bool:
        """分納ありかどうか"""
        return bool(self.installment_balance and self.installment_balance > 0)

    @property
    def welfare_limit(self) -> int | None:
        """備考欄から福祉上限額を読み取る。数値がなければNone。

        備考欄のパターン:
          - "80,000" or "110000" → そのまま数値
          - "9〇" or "11○" → 数字×10,000（○/〇 = 万の略記）
          - "9〇+初期1万" → 9×10,000 = 90,000（+以降は無視）
        """
        if not self.notes:
            return None
        import re
        notes_str = str(self.notes).strip()
        # まず「数字+○/〇」パターンを検出（○ = 万の略記）
        # 例: "9〇", "11○", "9〇+初期1万", "13"
        # +以降は無視
        base = notes_str.split("+")[0].strip()
        circle_match = re.match(r"^(\d+)\s*[○〇]", base)
        if circle_match:
            return int(circle_match.group(1)) * 10000

        # カンマ入り数値パターン（例: "80,000", "110,000"）
        cleaned = notes_str.replace(",", "").replace("，", "").strip()
        m = re.search(r"(\d+)", cleaned)
        if m:
            val = int(m.group(1))
            if val >= 10000:  # 万単位以上であれば上限額と判断
                return val
        return None

    @property
    def fixed_total(self) -> int:
        """固定費合計（D〜H列）"""
        return sum(v or 0 for v in [
            self.rent, self.management, self.common,
            self.water, self.utility
        ])


def _safe_int(val) -> int | None:
    """セル値を安全にintに変換。0は0として返す。None/空文字はNone。"""
    if val is None:
        return None
    if isinstance(val, str):
        val = val.strip()
        if not val:
            return None
        try:
            return int(float(val))
        except ValueError:
            return None
    if isinstance(val, (int, float)):
        # 数式の結果が0の場合（空行のSUM等）
        return int(val)
    return None


def _safe_str(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def read_resident_master(ws) -> list[ResidentMaster]:
    """入居者マスタシートを読む。

    セレーネ・ルネッサンス: 2列（A=居室, B=利用者名）
    パシフィック: 4列（A=居室, B=利用者名, C=請求摘要, D=対象期間）
    """
    residents = []
    for row in range(2, ws.max_row + 1):  # Row 1 はヘッダー
        room_raw = ws.cell(row=row, column=1).value
        name_raw = ws.cell(row=row, column=2).value
        if room_raw is None:
            continue
        room = normalize_room(room_raw)
        name = _safe_str(name_raw)
        if room:
            residents.append(ResidentMaster(room=room, name=name))
    return residents


def read_rx_sheet(ws) -> list[RxRow]:
    """RX.Xシートを読んで入居者データのリストを返す。

    Row 1: 日付
    Row 2: ヘッダー
    Row 3+: データ
    """
    rows = []
    for row_num in range(3, ws.max_row + 1):
        room_raw = ws.cell(row=row_num, column=COL["room"]).value
        name_raw = ws.cell(row=row_num, column=COL["name"]).value

        room = normalize_room(room_raw)
        name = _safe_str(name_raw)

        # 完全空行かチェック
        if not room and not name:
            # 合計行かもしれないので列Dをチェック
            rent_val = ws.cell(row=row_num, column=COL["rent"]).value
            if rent_val is not None and isinstance(rent_val, (int, float)) and rent_val > 0:
                # 合計行の可能性 — スキップ
                pass
            continue

        # 居室番号がなく、名前だけの場合もスキップ（合計行など）
        if not room:
            continue

        rx_row = RxRow(
            room=room,
            name=name,
            installment_balance=_safe_int(ws.cell(row=row_num, column=COL["installment_bal"]).value),
            rent=_safe_int(ws.cell(row=row_num, column=COL["rent"]).value),
            management=_safe_int(ws.cell(row=row_num, column=COL["management"]).value),
            common=_safe_int(ws.cell(row=row_num, column=COL["common"]).value),
            water=_safe_int(ws.cell(row=row_num, column=COL["water"]).value),
            utility=_safe_int(ws.cell(row=row_num, column=COL["utility"]).value),
            meal=_safe_int(ws.cell(row=row_num, column=COL["meal"]).value),
            adjustment=_safe_int(ws.cell(row=row_num, column=COL["adjustment"]).value),
            diaper=_safe_int(ws.cell(row=row_num, column=COL["diaper"]).value),
            daily_supplies=_safe_int(ws.cell(row=row_num, column=COL["daily_supplies"]).value),
            installment=_safe_int(ws.cell(row=row_num, column=COL["installment"]).value),
            office_fee=_safe_int(ws.cell(row=row_num, column=COL["office_fee"]).value),
            day_service=_safe_int(ws.cell(row=row_num, column=COL["day_service"]).value),
            welfare_equip=_safe_int(ws.cell(row=row_num, column=COL["welfare_equip"]).value),
            pharmacy=_safe_int(ws.cell(row=row_num, column=COL["pharmacy"]).value),
            doctor=_safe_int(ws.cell(row=row_num, column=COL["doctor"]).value),
            support=_safe_int(ws.cell(row=row_num, column=COL["support"]).value),
            other=_safe_int(ws.cell(row=row_num, column=COL["other"]).value),
            subtotal=_safe_int(ws.cell(row=row_num, column=COL["subtotal"]).value),
            care_burden=_safe_int(ws.cell(row=row_num, column=COL["care_burden"]).value),
            nurse_burden=_safe_int(ws.cell(row=row_num, column=COL["nurse_burden"]).value),
            total=_safe_int(ws.cell(row=row_num, column=COL["total"]).value),
            notes=_safe_str(ws.cell(row=row_num, column=COL["notes"]).value),
            invoice_status=_safe_str(ws.cell(row=row_num, column=COL["invoice_status"]).value),
            remaining_balance=_safe_int(ws.cell(row=row_num, column=COL["remaining_bal"]).value),
        )
        rows.append(rx_row)

    return rows


def _find_latest_rx_sheet(wb: openpyxl.Workbook, year: int, month: int) -> str | None:
    """対象月以前の最新RX.Xシートを探す。

    対象月のシートが見つからない場合、直前月から過去12ヶ月を遡って探す。
    """
    from utils.helpers import parse_reiwa_label

    # RX.X形式のシートを収集して年月でソート
    rx_sheets = []
    for name in wb.sheetnames:
        try:
            y, m = parse_reiwa_label(name)
            rx_sheets.append((y, m, name))
        except ValueError:
            continue

    if not rx_sheets:
        return None

    # 対象月以前のシートを新しい順にソート
    candidates = [(y, m, n) for y, m, n in rx_sheets if (y, m) < (year, month)]
    candidates.sort(key=lambda x: (x[0], x[1]), reverse=True)

    return candidates[0][2] if candidates else None


def read_master_file(
    filepath: str | Path,
    facility_config: dict,
    year: int,
    month: int,
    allow_fallback: bool = False,
) -> tuple[list[ResidentMaster], list[RxRow], bool]:
    """請求マスターファイルから入居者マスタとRX.Xデータを読む。

    Args:
        filepath: 請求マスターExcelファイルのパス
        facility_config: config.yaml の facilities.{拠点名} セクション
        year: 西暦年
        month: 月
        allow_fallback: Trueの場合、対象月シートがなければ直前月を使用

    Returns:
        (入居者マスタリスト, RX.Xデータリスト, is_new_month)
        is_new_month: Trueなら直前月のデータを使用（対象月シートが未存在）
    """
    wb = openpyxl.load_workbook(str(filepath), data_only=True)

    # 入居者マスタ
    resident_sheet = facility_config["resident_sheet"]
    if resident_sheet not in wb.sheetnames:
        wb.close()
        raise ValueError(
            f"Resident sheet '{resident_sheet}' not found. "
            f"Available: {wb.sheetnames}"
        )
    residents = read_resident_master(wb[resident_sheet])

    # RX.Xシート
    rx_label = to_reiwa_label(year, month)
    is_new_month = False

    if rx_label in wb.sheetnames:
        rx_rows = read_rx_sheet(wb[rx_label])
    elif allow_fallback:
        # 直前月のシートを探す
        fallback_label = _find_latest_rx_sheet(wb, year, month)
        if fallback_label is None:
            wb.close()
            raise ValueError(
                f"Sheet '{rx_label}' not found and no previous month sheet available. "
                f"Available: {wb.sheetnames}"
            )
        rx_rows = read_rx_sheet(wb[fallback_label])
        is_new_month = True
    else:
        wb.close()
        raise ValueError(
            f"Sheet '{rx_label}' not found. Available: {wb.sheetnames}"
        )

    wb.close()
    return residents, rx_rows, is_new_month


def read_all_facilities_masters(
    config: dict,
    year: int,
    month: int,
) -> dict[str, tuple[list[ResidentMaster], list[RxRow]]]:
    """全拠点の請求マスターを読む。

    config["input"] セクションからファイルパスを解決する。

    Returns:
        {拠点名: (入居者マスタ, RX.Xデータ)} の辞書
    """
    input_conf = config["input"]
    base_dir = Path(input_conf["base_dir"])
    result = {}

    for facility_name, fpath_conf in input_conf["facilities"].items():
        filepath = base_dir / fpath_conf["dir"] / fpath_conf["master"]
        if not filepath.exists():
            raise FileNotFoundError(
                f"Master file not found for {facility_name}: {filepath}"
            )
        fconf = config["facilities"][facility_name]
        residents, rx_rows, _ = read_master_file(filepath, fconf, year, month)
        result[facility_name] = (residents, rx_rows)

    return result
