"""ニック請求パーサー

ニック請求Excelファイルから、指定月の利用者ごとのオムツ・日用品データを読み取る。

構造:
  シート名: YYYYM 形式（例: '20261' = 2026年1月）
  Row 1: A1=月初日（datetime）
  Row 2: ヘッダー
  Row 3+: 利用者データ（1〜2行/人）
  列A: ステーション（結合あり）
  列B: 利用者ID（結合あり）
  列C: 氏名（結合あり）
  列D: セット種別（全角: Ａ,Ｂ,Ｃ,Ｄ,Ｅ,Ｆ,福,ふ,用）
  列E〜AI: 日1〜31（○=利用あり）
"""

from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

from utils.helpers import (
    col_letter_to_index,
    find_matching_sheet,
    get_sheet_name_candidates,
    nick_sheet_name,
)


# ○（全角丸）の文字
CIRCLE_CHAR = "\u25CB"  # ○ WHITE CIRCLE

# オムツ系セット（K列に集計されるもの）
DIAPER_SETS = {"Ａ", "Ｂ", "Ｃ", "Ｄ", "Ｅ", "Ｆ", "Ｇ",
               "a", "b", "c", "d"}
# 日用品系セット（L列に集計されるもの）
SUPPLY_SETS = {"用", "福", "ふ", "に"}

DAY_COL_START = col_letter_to_index("E")   # 5
DAY_COL_END = col_letter_to_index("AI")    # 35


@dataclass
class NickSetRecord:
    """1利用者の1セット分のデータ"""
    set_type: str         # セット種別（全角文字）
    use_days: list[int] = field(default_factory=list)  # 利用日リスト (1-indexed)
    day_count: int = 0    # 利用日数

    @property
    def is_diaper(self) -> bool:
        return self.set_type in DIAPER_SETS

    @property
    def is_supply(self) -> bool:
        return self.set_type in SUPPLY_SETS


@dataclass
class NickRecord:
    """1利用者のニック請求データ（複数セット可）"""
    station: str          # ステーション
    user_id: str          # 利用者ID
    name: str             # 氏名（姓 名、スペース付き）
    sets: list[NickSetRecord] = field(default_factory=list)

    @property
    def name_normalized(self) -> str:
        """スペースを除去した名前"""
        return self.name.replace(" ", "").replace("\u3000", "")

    def diaper_sets(self) -> list[NickSetRecord]:
        return [s for s in self.sets if s.is_diaper]

    def supply_sets(self) -> list[NickSetRecord]:
        return [s for s in self.sets if s.is_supply]


def _get_merged_value(ws, row: int, col: int):
    """結合セルの値を取得する。"""
    cell = ws.cell(row=row, column=col)
    for merge_range in ws.merged_cells.ranges:
        if cell.coordinate in merge_range:
            return ws.cell(row=merge_range.min_row, column=merge_range.min_col).value
    return cell.value


def _is_circle(val) -> bool:
    """セルの値が○（丸）かどうか判定"""
    if val is None:
        return False
    return str(val).strip() == CIRCLE_CHAR


def _read_use_days(ws, row: int, max_day: int = 31) -> tuple[list[int], int]:
    """1行分の利用日を読み取る。"""
    days = []
    count = 0
    for col in range(DAY_COL_START, DAY_COL_START + max_day):
        day_num = col - DAY_COL_START + 1
        val = ws.cell(row=row, column=col).value
        if _is_circle(val):
            days.append(day_num)
            count += 1
    return days, count


def _is_merged_with_above(ws, row: int, col: int) -> bool:
    """指定セルが上の行と結合されているかチェック"""
    cell = ws.cell(row=row, column=col)
    for merge_range in ws.merged_cells.ranges:
        if cell.coordinate in merge_range:
            return merge_range.min_row < row
    return False


def read_nick_sheet(ws, max_day: int = 31) -> list[NickRecord]:
    """ニック請求の1シートを解析する。

    Args:
        ws: openpyxlのワークシート
        max_day: 月の日数（1月=31, 2月=28or29, etc.）

    Returns:
        NickRecordのリスト
    """
    records = []
    data_start = 3
    max_row = ws.max_row

    current_record = None

    for row in range(data_start, max_row + 1):
        set_type_raw = ws.cell(row=row, column=4).value  # D列
        if set_type_raw is None:
            # セット種別が空なら空行（データ終了の可能性）
            if current_record is not None:
                records.append(current_record)
                current_record = None
            continue

        set_type = str(set_type_raw).strip()
        if not set_type:
            if current_record is not None:
                records.append(current_record)
                current_record = None
            continue

        # 日別フラグを読み取り
        use_days, day_count = _read_use_days(ws, row, max_day)
        set_record = NickSetRecord(
            set_type=set_type,
            use_days=use_days,
            day_count=day_count,
        )

        # A,B,C列が結合で上の行とつながっている場合 → 同一人物の追加セット
        is_continuation = _is_merged_with_above(ws, row, 3)  # C列で判定

        if is_continuation and current_record is not None:
            current_record.sets.append(set_record)
        else:
            # 新しい利用者
            if current_record is not None:
                records.append(current_record)

            station_raw = _get_merged_value(ws, row, 1)
            user_id_raw = _get_merged_value(ws, row, 2)
            name_raw = _get_merged_value(ws, row, 3)

            station = str(station_raw).strip() if station_raw else ""
            user_id = str(int(user_id_raw)) if isinstance(user_id_raw, (int, float)) else str(user_id_raw or "")
            name = str(name_raw).strip() if name_raw else ""

            current_record = NickRecord(
                station=station,
                user_id=user_id,
                name=name,
                sets=[set_record],
            )

    # 最後のレコードを追加
    if current_record is not None:
        records.append(current_record)

    return records


def read_nick_file(
    filepath: str | Path,
    year: int,
    month: int,
) -> list[NickRecord]:
    """ニック請求ファイルから指定月のデータを読み取る。

    Args:
        filepath: ニック請求Excelファイルのパス
        year: 西暦年
        month: 月

    Returns:
        NickRecordのリスト
    """
    import calendar
    max_day = calendar.monthrange(year, month)[1]

    wb = openpyxl.load_workbook(str(filepath), data_only=True)

    sheet_name = find_matching_sheet(wb.sheetnames, year, month)
    if sheet_name is None:
        candidates = get_sheet_name_candidates(year, month)
        wb.close()
        raise ValueError(
            f"Sheets {candidates} not found in {filepath}. "
            f"Available: {wb.sheetnames}"
        )

    ws = wb[sheet_name]
    records = read_nick_sheet(ws, max_day)
    wb.close()
    return records
