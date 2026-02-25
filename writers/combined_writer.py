"""合算明細書Excel生成モジュール

ええすまい＋ええかいご＋ええかんごの合計を一覧にした合算明細書を生成する。

テンプレート構造:
  A1:I1  "ご利用金額のお知らせ"
  A2     利用者名  H2=発行日
  A4     拠点名    F4=居室番号
  D7     合計金額（ええすまい+介護+看護）

  A10    "ええすまい"  C10=小計(W列)   E10=支払期限
  A11    "ええかいご"  C11=介護負担(X列) E11=支払期限
  A12    "ええかんご"  C12=看護負担(Y列) E12=支払期限

  Row 17~: 振込先情報
    住信SBIネット銀行（0038）
    法人第一支店（106）
    普通 2555848
    株式会社AA
"""

from datetime import date, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

from calc.billing import BillingResult
from utils.helpers import to_reiwa_label


def write_combined(
    billing: BillingResult,
    display_name: str,
    year: int,
    month: int,
    issue_date: date,
    output_path: Path,
) -> Path:
    """1入居者分の合算明細書Excelファイルを生成する。

    Args:
        billing: BillingResult
        display_name: 表示用拠点名
        year: 対象年
        month: 対象月
        issue_date: 発行日
        output_path: 出力先パス

    Returns:
        保存先Path
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "合算明細書"

    # フォント
    title_font = Font(name="游ゴシック", bold=True, size=14)
    name_font = Font(name="游ゴシック", bold=True, size=12)
    normal_font = Font(name="游ゴシック", size=10)
    header_font = Font(name="游ゴシック", bold=True, size=10)
    amount_font = Font(name="游ゴシック", bold=True, size=14)
    small_font = Font(name="游ゴシック", size=9)

    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    double_bottom = Border(bottom=Side(style="double"))

    # 支払期限: 発行日+20日（ええかんごは+25日）
    due_date_normal = issue_date + timedelta(days=20)
    due_date_nurse = issue_date + timedelta(days=25)

    # --- Row 1: タイトル ---
    ws.merge_cells("A1:I1")
    ws["A1"].value = "ご利用金額のお知らせ"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")

    # --- Row 2: 利用者名、発行日 ---
    ws.merge_cells("A2:D2")
    ws["A2"].value = billing.name
    ws["A2"].font = name_font
    ws["E2"].value = "様"
    ws["E2"].font = normal_font
    ws.merge_cells("H2:I2")
    ws["H2"].value = issue_date
    ws["H2"].number_format = "YYYY/MM/DD"
    ws["H2"].font = normal_font
    ws["H2"].alignment = Alignment(horizontal="right")

    # --- Row 4: 拠点名、居室番号 ---
    ws.merge_cells("A4:E4")
    ws["A4"].value = display_name
    ws["A4"].font = normal_font
    ws["F4"].value = billing.room
    ws["F4"].font = normal_font
    ws["G4"].value = "号室"
    ws["G4"].font = normal_font

    # --- Row 6: ヘッダー ---
    ws["A6"].value = "事業"
    ws["A6"].font = header_font
    ws["A6"].border = thin_border
    ws.merge_cells("C6:D6")
    ws["C6"].value = "ご請求金額"
    ws["C6"].font = header_font
    ws["C6"].border = thin_border
    ws["C6"].alignment = Alignment(horizontal="center")
    ws["D6"].border = thin_border
    ws.merge_cells("E6:F6")
    ws["E6"].value = "お支払い期限"
    ws["E6"].font = header_font
    ws["E6"].border = thin_border
    ws["E6"].alignment = Alignment(horizontal="center")
    ws["F6"].border = thin_border

    # --- Row 7: 合計金額 ---
    ws["A7"].value = "合計"
    ws["A7"].font = header_font
    ws.merge_cells("D7:E7")
    ws["D7"].value = billing.total
    ws["D7"].font = amount_font
    ws["D7"].number_format = "#,##0"
    ws["D7"].alignment = Alignment(horizontal="center")
    ws["D7"].border = double_bottom

    # --- Rows 10-12: 事業別内訳 ---
    items = [
        (10, "ええすまい", billing.subtotal, due_date_normal),
        (11, "ええかいご", billing.care_burden, due_date_normal),
        (12, "ええかんご", billing.nurse_burden, due_date_nurse),
    ]

    for row_num, label, amount, due in items:
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
        ws.cell(row=row_num, column=1).value = label
        ws.cell(row=row_num, column=1).font = normal_font
        ws.cell(row=row_num, column=1).border = thin_border
        ws.cell(row=row_num, column=2).border = thin_border

        ws.merge_cells(start_row=row_num, start_column=3, end_row=row_num, end_column=4)
        ws.cell(row=row_num, column=3).value = amount if amount else None
        ws.cell(row=row_num, column=3).font = normal_font
        ws.cell(row=row_num, column=3).number_format = "#,##0"
        ws.cell(row=row_num, column=3).alignment = Alignment(horizontal="right")
        ws.cell(row=row_num, column=3).border = thin_border
        ws.cell(row=row_num, column=4).border = thin_border

        ws.merge_cells(start_row=row_num, start_column=5, end_row=row_num, end_column=6)
        ws.cell(row=row_num, column=5).value = due if amount else None
        ws.cell(row=row_num, column=5).number_format = "YYYY/MM/DD"
        ws.cell(row=row_num, column=5).font = small_font
        ws.cell(row=row_num, column=5).alignment = Alignment(horizontal="center")
        ws.cell(row=row_num, column=5).border = thin_border
        ws.cell(row=row_num, column=6).border = thin_border

    # --- Row 14: 空行 ---

    # --- Row 15: お振込先 ---
    ws.merge_cells("A15:I15")
    ws["A15"].value = "＜お振込先＞"
    ws["A15"].font = header_font

    # --- Row 17-20: 振込先情報 ---
    bank_info = [
        "住信SBIネット銀行（0038）",
        "法人第一支店（106）",
        "普通　2555848",
        "株式会社AA",
    ]
    for i, text in enumerate(bank_info):
        row_num = 17 + i
        ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=6)
        ws.cell(row=row_num, column=2).value = text
        ws.cell(row=row_num, column=2).font = normal_font

    # --- 列幅設定 ---
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 8
    ws.column_dimensions["G"].width = 6
    ws.column_dimensions["H"].width = 8
    ws.column_dimensions["I"].width = 8

    # --- 印刷設定 ---
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.print_area = "A1:I22"

    # 保存
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    wb.close()

    return output_path


def write_all_combined(
    billings: list[BillingResult],
    display_name: str,
    year: int,
    month: int,
    issue_date: date,
    output_dir: Path,
) -> list[Path]:
    """全入居者分の合算明細書を生成する。

    Returns:
        生成されたファイルパスのリスト
    """
    reiwa_label = to_reiwa_label(year, month)
    files = []

    for billing in billings:
        if billing.is_vacant:
            continue
        # 介護・看護負担がある場合のみ合算明細を生成
        if not billing.care_burden and not billing.nurse_burden:
            continue

        filename = f"合算明細書_{billing.name}_{reiwa_label}.xlsx"
        filepath = output_dir / filename

        write_combined(
            billing=billing,
            display_name=display_name,
            year=year,
            month=month,
            issue_date=issue_date,
            output_path=filepath,
        )
        files.append(filepath)

    return files
