"""請求書Excel生成モジュール

既存テンプレートのレイアウトを再現して、入居者ごとの請求書Excelを生成する。

テンプレート構造（全拠点共通）:
  A1:K1  "ええすまい請求書" or "ええすまい御請求書"
  A2:D3  利用者名
  E2     "様"
  H2:I2  "発行日："
  J2:K2  発行日
  M2:N2  対象月（例: "R8.1"）
  M3     "家賃"  N3=翌月番号  O3="当月"  P3=当月番号  Q3="先月"  R3=前月番号
  A4:E4  拠点名
  F4     居室番号  G4="号室"
  H4:K5  "株式会社AA 代表取締役 大石佐智子"
  A6:B6  "ご請求金額"
  C6:G6  合計額
  H6:K6  住所

  A9="内容" C9="単価" E9="数量" G9="小計" I9="備考"
  Rows 10-26: 請求項目（家賃〜その他）
  G29: ご請求額合計
  I34: 税込合計
"""

from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter

from calc.billing import BillingResult
from utils.helpers import to_reiwa_label


# 請求項目定義: (行番号, ラベル, BillingResult属性名, 月区分: "next"/"current"/"prev")
BILLING_ITEMS = [
    (10, "家賃", "rent", "next"),
    (11, "管理費", "management", "next"),
    (12, "共益費", "common", "next"),
    (13, "水道料金", "water", "next"),
    (14, "定額光熱費", "utility", "next"),
    (15, "食事", "meal", "current"),
    (16, "調整額", "adjustment", "current"),
    (17, "オムツ", "diaper", "current"),
    (18, "日用品", "daily_supplies", "current"),
    (19, "分割支払い", "installment", "current"),
    (20, "事務手数料", "office_fee", "current"),
    (21, "デイサービス", "day_service", "prev"),
    (22, "福祉用具", "welfare_equip", "prev"),
    (23, "しろくま薬局", "pharmacy", "prev"),
    (24, "往診診療費", "doctor", "prev"),
    (25, "サポート費", "support", "prev"),
    (26, "その他", "other", "prev"),
]


def _get_month_numbers(year: int, month: int) -> tuple[int, int, int]:
    """対象月から、翌月番号・当月番号・前月番号を返す。"""
    current = month
    next_m = month + 1 if month < 12 else 1
    prev_m = month - 1 if month > 1 else 12
    return next_m, current, prev_m


def write_invoice(
    billing: BillingResult,
    facility_name: str,
    display_name: str,
    year: int,
    month: int,
    issue_date: date,
    output_path: Path,
) -> Path:
    """1入居者分の請求書Excelファイルを生成する。

    Args:
        billing: BillingResult
        facility_name: 拠点名（キー）
        display_name: 表示用拠点名（例: "マンションセレーネ（ええすまい）"）
        year: 対象年
        month: 対象月
        issue_date: 発行日
        output_path: 出力先パス

    Returns:
        保存先Path
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "すまい請求書"

    # フォント定義
    title_font = Font(name="游ゴシック", bold=True, size=16)
    name_font = Font(name="游ゴシック", bold=True, size=14)
    normal_font = Font(name="游ゴシック", size=10)
    header_font = Font(name="游ゴシック", bold=True, size=10)
    amount_font = Font(name="游ゴシック", bold=True, size=14)
    small_font = Font(name="游ゴシック", size=8)

    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    bottom_border = Border(bottom=Side(style="double"))

    reiwa_label = to_reiwa_label(year, month)
    next_m, current_m, prev_m = _get_month_numbers(year, month)

    # --- Row 1: タイトル ---
    ws.merge_cells("A1:K1")
    ws["A1"].value = "ええすまい請求書"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")

    # --- Row 2-3: 利用者名、発行日、対象月 ---
    ws.merge_cells("A2:D3")
    ws["A2"].value = billing.name
    ws["A2"].font = name_font
    ws["E2"].value = "様"
    ws["E2"].font = normal_font

    ws.merge_cells("H2:I2")
    ws["H2"].value = "発行日："
    ws["H2"].font = normal_font
    ws["H2"].alignment = Alignment(horizontal="right")

    ws.merge_cells("J2:K2")
    ws["J2"].value = issue_date
    ws["J2"].number_format = "YYYY/MM/DD"
    ws["J2"].font = normal_font

    ws.merge_cells("M2:N2")
    ws["M2"].value = reiwa_label
    ws["M2"].font = small_font

    ws["M3"].value = "家賃"
    ws["M3"].font = small_font
    ws["N3"].value = next_m
    ws["N3"].font = small_font
    ws["O3"].value = "当月"
    ws["O3"].font = small_font
    ws["P3"].value = current_m
    ws["P3"].font = small_font
    ws["Q3"].value = "先月"
    ws["Q3"].font = small_font
    ws["R3"].value = prev_m
    ws["R3"].font = small_font

    # --- Row 4-5: 拠点名、居室番号、会社情報 ---
    ws.merge_cells("A4:E4")
    ws["A4"].value = display_name
    ws["A4"].font = normal_font

    ws["F4"].value = billing.room
    ws["F4"].font = normal_font
    ws["G4"].value = "号室"
    ws["G4"].font = normal_font

    ws.merge_cells("H4:K5")
    ws["H4"].value = "株式会社AA\n代表取締役 大石佐智子"
    ws["H4"].font = normal_font
    ws["H4"].alignment = Alignment(wrap_text=True)

    # --- Row 6: ご請求金額、住所 ---
    ws.merge_cells("A6:B6")
    ws["A6"].value = "ご請求金額"
    ws["A6"].font = header_font

    ws.merge_cells("C6:G6")
    ws["C6"].value = billing.total
    ws["C6"].font = amount_font
    ws["C6"].number_format = "#,##0"
    ws["C6"].border = bottom_border
    ws["C6"].alignment = Alignment(horizontal="center")

    ws.merge_cells("H6:K6")
    ws["H6"].value = "大阪市西淀川区柏里2-9-3-102"
    ws["H6"].font = small_font

    # --- Row 8: 空行 ---

    # --- Row 9: 項目ヘッダー ---
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    headers = [(1, "内容"), (3, "単価"), (5, "数量"), (7, "小計"), (9, "備考")]
    for col_idx, text in headers:
        ws.merge_cells(
            start_row=9, start_column=col_idx,
            end_row=9, end_column=col_idx + 1
        )
        cell = ws.cell(row=9, column=col_idx)
        cell.value = text
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
        # 右セルにも罫線
        ws.cell(row=9, column=col_idx + 1).border = thin_border

    # --- Rows 10-26: 請求項目 ---
    month_label_map = {"next": next_m, "current": current_m, "prev": prev_m}

    for row_num, label, attr, period in BILLING_ITEMS:
        value = getattr(billing, attr, 0) or 0

        # A-B列: 内容
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
        ws.cell(row=row_num, column=1).value = label
        ws.cell(row=row_num, column=1).font = normal_font
        ws.cell(row=row_num, column=1).border = thin_border
        ws.cell(row=row_num, column=2).border = thin_border

        # C-D列: 単価
        ws.merge_cells(start_row=row_num, start_column=3, end_row=row_num, end_column=4)
        ws.cell(row=row_num, column=3).value = value if value != 0 else None
        ws.cell(row=row_num, column=3).font = normal_font
        ws.cell(row=row_num, column=3).number_format = "#,##0"
        ws.cell(row=row_num, column=3).alignment = Alignment(horizontal="right")
        ws.cell(row=row_num, column=3).border = thin_border
        ws.cell(row=row_num, column=4).border = thin_border

        # E-F列: 数量
        ws.merge_cells(start_row=row_num, start_column=5, end_row=row_num, end_column=6)
        if value and value != 0:
            ws.cell(row=row_num, column=5).value = "1か月" if row_num <= 14 else 1
        ws.cell(row=row_num, column=5).font = normal_font
        ws.cell(row=row_num, column=5).alignment = Alignment(horizontal="center")
        ws.cell(row=row_num, column=5).border = thin_border
        ws.cell(row=row_num, column=6).border = thin_border

        # G-H列: 小計
        ws.merge_cells(start_row=row_num, start_column=7, end_row=row_num, end_column=8)
        ws.cell(row=row_num, column=7).value = value if value != 0 else None
        ws.cell(row=row_num, column=7).font = normal_font
        ws.cell(row=row_num, column=7).number_format = "#,##0"
        ws.cell(row=row_num, column=7).alignment = Alignment(horizontal="right")
        ws.cell(row=row_num, column=7).border = thin_border
        ws.cell(row=row_num, column=8).border = thin_border

        # I-K列: 備考（月番号）
        ws.merge_cells(start_row=row_num, start_column=9, end_row=row_num, end_column=11)
        ws.cell(row=row_num, column=9).value = month_label_map[period]
        ws.cell(row=row_num, column=9).font = small_font
        ws.cell(row=row_num, column=9).alignment = Alignment(horizontal="center")
        ws.cell(row=row_num, column=9).border = thin_border
        ws.cell(row=row_num, column=10).border = thin_border
        ws.cell(row=row_num, column=11).border = thin_border

    # --- Rows 27-28: 空行(罫線のみ) ---
    for row_num in range(27, 29):
        for col_idx in range(1, 12):
            ws.cell(row=row_num, column=col_idx).border = thin_border

    # --- Row 29: ご請求額合計 ---
    ws.merge_cells("A29:F29")
    ws["A29"].value = "ご請求額合計"
    ws["A29"].font = header_font
    ws["A29"].alignment = Alignment(horizontal="right")

    ws.merge_cells("G29:H29")
    ws["G29"].value = f"=SUM(G10:H28)"
    ws["G29"].font = amount_font
    ws["G29"].number_format = "#,##0"
    ws["G29"].alignment = Alignment(horizontal="right")
    ws["G29"].border = Border(bottom=Side(style="double"))

    # --- Row 30: 消費税注記 ---
    ws.merge_cells("A30:K30")
    ws["A30"].value = "※家賃、共益費、管理費、光熱費は消費税不要"
    ws["A30"].font = small_font

    # --- Row 32-33: 消費税計算（全項目消費税0） ---
    ws.merge_cells("A32:B32")
    ws["A32"].value = "8%対象 合計"
    ws["A32"].font = small_font
    ws.merge_cells("C32:D32")
    ws["C32"].value = 0
    ws["C32"].number_format = "#,##0"
    ws.merge_cells("G32:H32")
    ws["G32"].value = "消費税（8%）"
    ws["G32"].font = small_font
    ws.merge_cells("I32:K32")
    ws["I32"].value = 0
    ws["I32"].number_format = "#,##0"

    ws.merge_cells("A33:B33")
    ws["A33"].value = "10%対象 合計"
    ws["A33"].font = small_font
    ws.merge_cells("C33:D33")
    ws["C33"].value = 0
    ws["C33"].number_format = "#,##0"
    ws.merge_cells("G33:H33")
    ws["G33"].value = "消費税（10%）"
    ws["G33"].font = small_font
    ws.merge_cells("I33:K33")
    ws["I33"].value = 0
    ws["I33"].number_format = "#,##0"

    # --- Row 34: 税込合計 ---
    ws.merge_cells("G34:H34")
    ws["G34"].value = "税込合計"
    ws["G34"].font = header_font
    ws.merge_cells("I34:K34")
    ws["I34"].value = "=G29"
    ws["I34"].font = amount_font
    ws["I34"].number_format = "#,##0"
    ws["I34"].border = Border(bottom=Side(style="double"))

    # --- 列幅設定 ---
    col_widths = {"A": 6, "B": 8, "C": 8, "D": 6, "E": 6, "F": 4,
                  "G": 8, "H": 6, "I": 5, "J": 8, "K": 5}
    for letter, width in col_widths.items():
        ws.column_dimensions[letter].width = width

    # --- 印刷設定 ---
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.print_area = "A1:K34"

    # 保存
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    wb.close()

    return output_path


def write_all_invoices(
    billings: list[BillingResult],
    facility_name: str,
    display_name: str,
    year: int,
    month: int,
    issue_date: date,
    output_dir: Path,
) -> list[Path]:
    """全入居者分の請求書を生成する。

    Args:
        billings: BillingResultリスト
        facility_name: 拠点キー名
        display_name: 表示用拠点名
        year: 対象年
        month: 対象月
        issue_date: 発行日
        output_dir: 出力ディレクトリ（拠点ごとのサブディレクトリ）

    Returns:
        生成されたファイルパスのリスト
    """
    reiwa_label = to_reiwa_label(year, month)
    files = []

    for billing in billings:
        if billing.is_vacant:
            continue

        filename = f"請求書_{billing.room}_{billing.name}_{reiwa_label}.xlsx"
        filepath = output_dir / filename

        write_invoice(
            billing=billing,
            facility_name=facility_name,
            display_name=display_name,
            year=year,
            month=month,
            issue_date=issue_date,
            output_path=filepath,
        )
        files.append(filepath)

    return files
