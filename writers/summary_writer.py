"""RX.X集約シート書込モジュール

請求マスターの既存ファイルに新しいRX.Xシートを作成（or更新）し、
BillingResultのデータを書き込む。

列構造（新フォーマット R7.8以降）:
  A:居室 B:利用者名 C:分納残 D:家賃 E:管理費 F:共益費 G:水道料金
  H:定額光熱費 I:食事 J:調整額 K:オムツ L:日用品 M:分割支払い
  N:事務手数料 O:デイサービス P:福祉用具 Q:しろくま薬局
  R:往診診療費 S:サポート費 T:その他 U-V:(予備)
  W:計 X:一部負担(介護) Y:一部負担(看護) Z:合計
  AA:備考 AB:請求書 AC:分割残高
"""

import calendar
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from calc.billing import BillingResult
from utils.helpers import to_reiwa_label


# 列インデックス（1始まり）
COL_MAP = {
    "room": 1,            # A
    "name": 2,            # B
    "installment_bal": 3, # C
    "rent": 4,            # D
    "management": 5,      # E
    "common": 6,          # F
    "water": 7,           # G
    "utility": 8,         # H
    "meal": 9,            # I
    "adjustment": 10,     # J
    "diaper": 11,         # K
    "daily_supplies": 12, # L
    "installment": 13,    # M
    "office_fee": 14,     # N
    "day_service": 15,    # O
    "welfare_equip": 16,  # P
    "pharmacy": 17,       # Q
    "doctor": 18,         # R
    "support": 19,        # S
    "other": 20,          # T
    "subtotal": 23,       # W
    "care_burden": 24,    # X
    "nurse_burden": 25,   # Y
    "total": 26,          # Z
    "notes": 27,          # AA
    "invoice_status": 28, # AB
    "remaining_bal": 29,  # AC
}

HEADERS = [
    (1, "居室"), (2, "利用者名"), (3, "分納残"), (4, "家賃"), (5, "管理費"),
    (6, "共益費"), (7, "水道料金"), (8, "定額光熱費"), (9, "食事"), (10, "調整額"),
    (11, "オムツ"), (12, "日用品"), (13, "分割支払い"), (14, "事務手数料"),
    (15, "デイサービス"), (16, "福祉用具"), (17, "しろくま薬局"),
    (18, "往診診療費"), (19, "サポート費"), (20, "その他"),
    (21, None), (22, None),
    (23, "計"), (24, "一部負担(介護)"), (25, "一部負担(看護)"),
    (26, "合計"), (27, "備考"), (28, "請求書"), (29, "分割残高"),
]


def _write_cell(ws, row: int, col: int, value, is_number: bool = False):
    """セルに値を書き込む。0やNoneは空白にする（一部列を除く）。"""
    cell = ws.cell(row=row, column=col)
    if value is None or (is_number and value == 0):
        cell.value = None
    else:
        cell.value = value


def write_summary_sheet(
    wb: openpyxl.Workbook,
    sheet_name: str,
    billings: list[BillingResult],
    year: int,
    month: int,
) -> None:
    """ワークブックにRX.Xシートを作成/更新して請求データを書き込む。

    既存のシートがある場合は削除して再作成する。

    Args:
        wb: openpyxl Workbook
        sheet_name: シート名（例: "R8.1"）
        billings: BillingResultのリスト
        year: 西暦年
        month: 月
    """
    # 既存シートがあれば削除
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(title=sheet_name, index=0)

    # Row 1: 対象月末日
    last_day = calendar.monthrange(year, month)[1]
    month_end = date(year, month, last_day)
    ws.merge_cells("A1:B1")
    ws.cell(row=1, column=1, value=month_end)
    ws.cell(row=1, column=1).number_format = "YYYY-MM-DD"
    ws.cell(row=1, column=3, value="請求")

    # Row 2: ヘッダー
    header_font = Font(bold=True, size=9)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, header_text in HEADERS:
        cell = ws.cell(row=2, column=col_idx)
        cell.value = header_text
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Row 3+: データ行
    data_font = Font(size=9)
    for i, billing in enumerate(billings):
        row_num = 3 + i
        ws.cell(row=row_num, column=COL_MAP["room"], value=billing.room)
        ws.cell(row=row_num, column=COL_MAP["name"], value=billing.name if billing.name else None)

        # 分納残（0以外を表示）
        _write_cell(ws, row_num, COL_MAP["installment_bal"],
                     billing.installment_balance if billing.installment_balance > 0 else None)

        # 固定費
        ws.cell(row=row_num, column=COL_MAP["rent"], value=billing.rent or None)
        ws.cell(row=row_num, column=COL_MAP["management"], value=billing.management or None)
        ws.cell(row=row_num, column=COL_MAP["common"], value=billing.common or None)
        ws.cell(row=row_num, column=COL_MAP["water"], value=billing.water or None)
        ws.cell(row=row_num, column=COL_MAP["utility"], value=billing.utility or None)

        # 食事
        ws.cell(row=row_num, column=COL_MAP["meal"], value=billing.meal or None)
        # 調整額（マイナス値も書く）
        ws.cell(row=row_num, column=COL_MAP["adjustment"],
                value=billing.adjustment if billing.adjustment != 0 else None)
        # オムツ・日用品
        ws.cell(row=row_num, column=COL_MAP["diaper"], value=billing.diaper or None)
        ws.cell(row=row_num, column=COL_MAP["daily_supplies"], value=billing.daily_supplies or None)

        # 分割支払い
        ws.cell(row=row_num, column=COL_MAP["installment"], value=billing.installment or None)
        # 手入力項目
        ws.cell(row=row_num, column=COL_MAP["office_fee"], value=billing.office_fee or None)
        ws.cell(row=row_num, column=COL_MAP["day_service"], value=billing.day_service or None)
        ws.cell(row=row_num, column=COL_MAP["welfare_equip"], value=billing.welfare_equip or None)
        ws.cell(row=row_num, column=COL_MAP["pharmacy"], value=billing.pharmacy or None)
        ws.cell(row=row_num, column=COL_MAP["doctor"], value=billing.doctor or None)
        ws.cell(row=row_num, column=COL_MAP["support"], value=billing.support or None)
        ws.cell(row=row_num, column=COL_MAP["other"], value=billing.other or None)

        # W列: 計（SUM数式）
        ws.cell(row=row_num, column=COL_MAP["subtotal"],
                value=f"=SUM(D{row_num}:V{row_num})")

        # X,Y列: 介護・看護負担
        ws.cell(row=row_num, column=COL_MAP["care_burden"],
                value=billing.care_burden or None)
        ws.cell(row=row_num, column=COL_MAP["nurse_burden"],
                value=billing.nurse_burden or None)

        # Z列: 合計（SUM数式）
        ws.cell(row=row_num, column=COL_MAP["total"],
                value=f"=SUM(W{row_num}:Y{row_num})")

        # AA: 備考
        ws.cell(row=row_num, column=COL_MAP["notes"], value=billing.notes or None)

        # AC: 分割残高（数式）
        ws.cell(row=row_num, column=COL_MAP["remaining_bal"],
                value=f"=C{row_num}-M{row_num}")

        # 罫線適用
        for col_idx in range(1, 30):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.border = thin_border
            cell.font = data_font

    # 合計行
    total_row = 3 + len(billings) + 1
    ws.cell(row=total_row, column=1, value="合計")
    ws.cell(row=total_row, column=1).font = Font(bold=True, size=9)
    # 数値列にSUM数式を設定
    for col_idx in [4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 23, 24, 25, 26]:
        col_letter = get_column_letter(col_idx)
        ws.cell(row=total_row, column=col_idx,
                value=f"=SUM({col_letter}3:{col_letter}{total_row - 2})")
        ws.cell(row=total_row, column=col_idx).font = Font(bold=True, size=9)
        ws.cell(row=total_row, column=col_idx).border = thin_border

    # 列幅調整
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 14
    for col_idx in range(3, 30):
        ws.column_dimensions[get_column_letter(col_idx)].width = 10


def write_summary_to_file(
    master_filepath: str | Path,
    billings: list[BillingResult],
    year: int,
    month: int,
    output_filepath: str | Path | None = None,
) -> Path:
    """請求マスターファイルのRX.Xシートを更新して保存する。

    Args:
        master_filepath: 元の請求マスターファイルパス
        billings: BillingResultリスト
        year: 西暦年
        month: 月
        output_filepath: 出力先（Noneなら元ファイルと同じ名前で出力ディレクトリに保存）

    Returns:
        保存先のPath
    """
    sheet_name = to_reiwa_label(year, month)
    wb = openpyxl.load_workbook(str(master_filepath))

    write_summary_sheet(wb, sheet_name, billings, year, month)

    if output_filepath is None:
        output_filepath = master_filepath
    output_path = Path(output_filepath)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    wb.close()

    return output_path
